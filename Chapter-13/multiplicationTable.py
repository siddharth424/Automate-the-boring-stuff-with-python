import openpyxl
import sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
wb=openpyxl.Workbook()
n=int(sys.argv[1])
wb.create_sheet(index=0,title='First Sheeet')
sheet=wb.active
bold=Font(bold=True)
for i in range(2,n+2):
    sheet[get_column_letter(i)+str(1)]=i-1
    sheet['A'+str(i)]=i-1
    sheet[get_column_letter(i)+str(1)].font=bold
    sheet['A'+str(i)].font=bold
    

for row in range(2,n+2):
    for cell in range(2,n+2):
        sheet[get_column_letter(cell)+str(row)]=(row-1)*(cell-1)
wb.save('multi_table.xlsx')
        
        
